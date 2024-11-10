import openpyxl

# Criando o arquivo Excel
wb = openpyxl.Workbook()

# Planilha 1: Empregado
sheet_empregado = wb.create_sheet(title="Empregado")
sheet_empregado.append(["mss (chave)", "pnome", "mnome", "snome", "sexo", "datanasc", "salario", "endereco", "numero-dep (trabalha em)", "nss-supervisor (supervisiona)"])
sheet_empregado.append([1001, "João", "Carlos", "Silva", "M", "1990-05-15", 3500, "Rua A, 123", 1, 5001])
sheet_empregado.append([1002, "Maria", "José", "Souza", "F", "1992-08-22", 3200, "Rua B, 456", 2, 5002])

# Planilha 2: Departamento
sheet_departamento = wb.create_sheet(title="Departamento")
sheet_departamento.append(["numero (chave)", "nome", "nro_emp", "nss_emp (gerencia)", "datainicio (gerencia)"])
sheet_departamento.append([1, "TI", 5, 1001, "2020-01-10"])
sheet_departamento.append([2, "RH", 3, 1002, "2021-02-15"])

# Planilha 3: Projeto
sheet_projeto = wb.create_sheet(title="Projeto")
sheet_projeto.append(["numero (chave)", "nome", "localizacao", "numero-dep (controle)"])
sheet_projeto.append([101, "Projeto A", "Localizacao1", 1])
sheet_projeto.append([102, "Projeto B", "Localizacao2", 2])

# Planilha 4: Depende
sheet_depende = wb.create_sheet(title="Depende")
sheet_depende.append(["nss-emp (chave)", "nome (chave)", "sexo", "datanasc", "tiporel"])
sheet_depende.append([1001, "Maria Silva", "F", "2015-07-20", "Filha"])
sheet_depende.append([1002, "João Souza", "M", "2018-11-10", "Filho"])

# Planilha 5: Trabalha-em
sheet_trabalha_em = wb.create_sheet(title="Trabalha-em")
sheet_trabalha_em.append(["nss-emp (chave)", "numero-proj (chave)", "horas"])
sheet_trabalha_em.append([1001, 101, 40])
sheet_trabalha_em.append([1002, 102, 35])

# Planilha 6: Localizacao
sheet_localizacao = wb.create_sheet(title="Localizacao")
sheet_localizacao.append(["localizacao (chave)", "numero-dep (chave)"])
sheet_localizacao.append(["Localizacao1", 1])
sheet_localizacao.append(["Localizacao2", 2])

# Salvando o arquivo
wb.save("empresa_banco_dados.xlsx")

# Carregar o arquivo Excel
wb = openpyxl.load_workbook("empresa_banco_dados.xlsx")

# Listar os nomes das planilhas
print("Planilhas no arquivo:")
for sheet in wb.sheetnames:
    print(sheet)
